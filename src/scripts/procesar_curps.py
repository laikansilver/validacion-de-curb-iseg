#!/usr/bin/env python3
"""
Procesador de múltiples CURPs desde archivo CSV o XLSX
Valida cada CURP y indica cuáles son válidos y cuáles no
Usando procesamiento paralelo para mayor velocidad
"""

import csv
import sys
import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# Agregar ruta para imports
sys.path.insert(0, str(Path(__file__).parent.parent / "core"))
from cliente_cdc import CirculoDeCredito

# Cargar configuración
config_path = Path("config/config_cdc.json")
with open(config_path, 'r') as f:
    config = json.load(f)

cdc_config = config.get("cdc", {})
client = CirculoDeCredito(
    api_key=cdc_config.get("api_key"),
    private_key_b64=cdc_config.get("private_key_b64"),
    base_url=cdc_config.get("base_url")
)

# Lock para thread-safety
print_lock = threading.Lock()


def validar_curp_individual(numero: int, curp: str) -> dict:
    """Valida un CURP individual (para uso en ThreadPoolExecutor)"""
    try:
        exitoso, respuesta = client.validate_curp(curp)
        
        if exitoso:
            datos = client.extract_data(respuesta)
            if datos and datos.get("curp"):
                return {
                    "numero": numero,
                    "curp": curp,
                    "estado": "REGISTRADO",
                    "datos": datos,
                    "respuesta_completa": respuesta
                }
            else:
                return {
                    "numero": numero,
                    "curp": curp,
                    "estado": "NO_ENCONTRADO",
                    "razon": "CURP no registrado ante el gobierno"
                }
        else:
            return {
                "numero": numero,
                "curp": curp,
                "estado": "ERROR",
                "error": respuesta.get("error", str(respuesta))
            }
    except Exception as e:
        return {
            "numero": numero,
            "curp": curp,
            "estado": "ERROR",
            "error": str(e)
        }


def validar_lote_curps(archivo_entrada: str) -> dict:
    """
    Procesa un CSV o XLSX con múltiples CURPs y valida cada uno EN PARALELO
    
    Espera columna 'curp' en el archivo
    """
    print("=" * 80)
    print("📊 PROCESADOR DE MÚLTIPLES CURPs (PARALELO)")
    print("=" * 80)
    print()
    
    archivo = Path(archivo_entrada)
    if not archivo.exists():
        print(f"❌ Error: No se encontró el archivo '{archivo_entrada}'")
        return None
    
    print(f"📂 Archivo: {archivo.name}")
    print()
    
    # Detectar tipo de archivo
    es_excel = archivo.suffix.lower() == ".xlsx"
    
    # Leer todos los CURPs primero
    curps_list = []
    
    try:
        if es_excel:
            workbook = load_workbook(archivo)
            worksheet = workbook.active
            
            # Obtener encabezados
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value)
            
            if 'curp' not in [h.lower() if h else None for h in headers]:
                print(f"❌ Error: El Excel debe tener una columna 'curp'")
                print(f"   Columnas encontradas: {', '.join([str(h) for h in headers if h])}")
                return None
            
            curp_index = next(i for i, h in enumerate(headers) if h.lower() == 'curp')
            
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 1):
                curp = row[curp_index] if curp_index < len(row) else None
                if curp:
                    curps_list.append((idx, str(curp).strip().upper()))
        
        else:
            with open(archivo, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                
                if reader.fieldnames and 'curp' not in reader.fieldnames:
                    print(f"❌ Error: El CSV debe tener una columna 'curp'")
                    print(f"   Columnas encontradas: {', '.join(reader.fieldnames)}")
                    return None
                
                for idx, row in enumerate(reader, 1):
                    curp = row.get('curp', '').strip().upper()
                    if curp:
                        curps_list.append((idx, curp))
        
        print(f"📋 {len(curps_list)} CURPs a validar")
        print()
        
        # Procesar en paralelo
        validos = []
        invalidos = []
        errores = []
        procesados = 0
        
        # Usar máximo 5 threads concurrentes para no sobrecargar la API
        max_workers = 5
        
        print(f"⚙️  Iniciando validación paralela ({max_workers} threads)...")
        print()
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(validar_curp_individual, num, curp): (num, curp) 
                      for num, curp in curps_list}
            
            for future in as_completed(futures):
                num, curp = futures[future]
                resultado = future.result()
                procesados += 1
                
                # Mostrar progreso
                estado_icono = "✅" if resultado["estado"] == "REGISTRADO" else "❌" if resultado["estado"] == "NO_ENCONTRADO" else "⚠️"
                print(f"🔍 [{resultado['numero']}] {curp}... {estado_icono} {resultado['estado']}")
                
                # Clasificar resultado
                if resultado["estado"] == "REGISTRADO":
                    validos.append({
                        "numero": resultado["numero"],
                        "curp_solicitado": curp,
                        "datos": resultado["datos"],
                        "respuesta_completa": resultado["respuesta_completa"]
                    })
                elif resultado["estado"] == "NO_ENCONTRADO":
                    invalidos.append({
                        "numero": resultado["numero"],
                        "curp": curp,
                        "razon": resultado.get("razon", "CURP no registrado")
                    })
                else:
                    errores.append({
                        "numero": resultado["numero"],
                        "curp": curp,
                        "error": resultado.get("error", "Error desconocido")
                    })
        
        total = len(curps_list)
        
        print()
        print("=" * 80)
        print("📋 RESUMEN DE RESULTADOS")
        print("=" * 80)
        print()
        
        print(f"Total procesados:       {total}")
        print(f"✅ Registrados:         {len(validos)}")
        print(f"❌ No registrados:      {len(invalidos)}")
        print(f"⚠️  Errores:            {len(errores)}")
        print()
        
        return {
            "total": total,
            "validos": validos,
            "invalidos": invalidos,
            "errores": errores
        }
        
    except Exception as e:
        print(f"❌ Error inesperado: {e}")
        import traceback
        traceback.print_exc()
        return None


def guardar_resultados(resultados: dict, nombre_salida: str = None):
    """Guarda los resultados en archivos"""
    if not resultados:
        return
    
    # Generar nombre de archivo si no se proporciona
    if not nombre_salida:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_salida = f"validaciones_{timestamp}"
    
    # Crear directorio de resultados
    resultado_dir = Path("data/resultados")
    resultado_dir.mkdir(parents=True, exist_ok=True)
    
    # Guardar JSON completo
    archivo_json = resultado_dir / f"{nombre_salida}.json"
    with open(archivo_json, 'w', encoding='utf-8') as f:
        # Separar datos para no incluir respuesta completa en JSON
        datos_guardar = {
            "total": resultados["total"],
            "resumen": {
                "validos": len(resultados["validos"]),
                "invalidos": len(resultados["invalidos"]),
                "errores": len(resultados["errores"])
            },
            "validos": [
                {
                    "numero": v["numero"],
                    "curp_solicitado": v["curp_solicitado"],
                    "datos": v["datos"]
                }
                for v in resultados["validos"]
            ],
            "invalidos": resultados["invalidos"],
            "errores": resultados["errores"]
        }
        json.dump(datos_guardar, f, indent=2, ensure_ascii=False)
    
    # Guardar CSV con resultados
    archivo_csv = resultado_dir / f"{nombre_salida}.csv"
    with open(archivo_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        
        # Encabezado
        writer.writerow([
            "Número",
            "CURP Solicitado",
            "Estado",
            "CURP Retornado",
            "Nombres",
            "Apellido Paterno",
            "Apellido Materno",
            "Sexo",
            "Fecha Nacimiento",
            "Entidad Nacimiento",
            "Nacionalidad",
            "Detalles"
        ])
        
        # Válidos
        for v in resultados["validos"]:
            datos = v["datos"]
            writer.writerow([
                v["numero"],
                v["curp_solicitado"],
                "✅ REGISTRADO",
                datos.get("curp"),
                datos.get("nombres"),
                datos.get("apellidoPaterno"),
                datos.get("apellidoMaterno"),
                datos.get("sexo"),
                datos.get("fechaNacimiento"),
                datos.get("entidadNacimiento"),
                datos.get("nacionalidad"),
                ""
            ])
        
        # Inválidos
        for inv in resultados["invalidos"]:
            writer.writerow([
                inv["numero"],
                inv["curp"],
                "❌ NO REGISTRADO",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                inv.get("razon")
            ])
        
        # Errores
        for err in resultados["errores"]:
            writer.writerow([
                err["numero"],
                err["curp"],
                "⚠️ ERROR",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                err.get("error")
            ])


def main():
    """Función principal"""
    if len(sys.argv) < 2:
        print("Uso: python procesar_curps.py <archivo.csv|xlsx> [nombre_salida]")
        print()
        print("Ejemplo:")
        print("  python procesar_curps.py data/curps.csv")
        print("  python procesar_curps.py data/curps.xlsx resultados_mayo")
        print()
        print("Formatos soportados:")
        print("  - CSV (columna 'curp')")
        print("  - XLSX (columna 'curp')")
        print()
        print("Ejemplo de CSV:")
        print("  curp")
        print("  MERE020526HMNRZDA2")
        print("  PERE030615HMNMRA05")
        print("  ...")
        sys.exit(1)
    
    archivo = sys.argv[1]
    nombre_salida = sys.argv[2] if len(sys.argv) > 2 else None
    
    # Procesar
    resultados = validar_lote_curps(archivo)
    
    if resultados:
        guardar_resultados(resultados, nombre_salida)


if __name__ == "__main__":
    main()
